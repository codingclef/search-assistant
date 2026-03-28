"""
Tests for modules/excel_writer.py

Verifies sheet structure, content routing, and formatting.
"""
import io
from datetime import datetime

import openpyxl
import pytest

from modules.excel_writer import create_excel


# ── helpers ───────────────────────────────────────────────────────────────────

def _article(title="테스트 기사", category="부정기사", reason="테스트 이유"):
    return {
        "keyword": "테스트키워드",
        "title": title,
        "link": "https://example.com/news/1",
        "published_at": datetime(2024, 3, 25, 9, 30),
        "search_engine": "네이버",
        "source": "테스트신문",
        "description": "테스트 설명",
        "category": category,
        "reason": reason,
    }


def _load(result: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(result))


# ── basic output ──────────────────────────────────────────────────────────────

def test_returns_bytes():
    assert isinstance(create_excel([], [], lang="ko"), bytes)


def test_output_is_non_empty():
    assert len(create_excel([], [], lang="ko")) > 0


def test_output_is_valid_xlsx():
    wb = _load(create_excel([], [], lang="ko"))
    assert wb is not None


# ── sheet names ───────────────────────────────────────────────────────────────

def test_ko_required_sheets_exist():
    wb = _load(create_excel([], [], lang="ko"))
    assert "일람" in wb.sheetnames
    assert "보류" in wb.sheetnames
    assert "해당없음" in wb.sheetnames


def test_ja_required_sheets_exist():
    wb = _load(create_excel([], [], lang="ja"))
    assert "一覧" in wb.sheetnames
    assert "保留" in wb.sheetnames
    assert "該当なし" in wb.sheetnames


def test_custom_category_sheet_is_created():
    wb = _load(create_excel([_article()], ["부정기사"], lang="ko"))
    assert "부정기사" in wb.sheetnames


def test_multiple_custom_category_sheets():
    articles = [
        _article(category="부정기사"),
        _article(category="중요기사"),
    ]
    wb = _load(create_excel(articles, ["부정기사", "중요기사"], lang="ko"))
    assert "부정기사" in wb.sheetnames
    assert "중요기사" in wb.sheetnames


# ── sheet order ───────────────────────────────────────────────────────────────

def test_sheet_order_ilam_first_then_categories_then_holdup_na():
    articles = [_article(category="부정기사")]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    names = wb.sheetnames
    assert names[0] == "일람"
    assert names[-1] == "해당없음"
    assert names[-2] == "보류"
    assert "부정기사" in names[1:-2]


# ── ilam sheet content ────────────────────────────────────────────────────────

def test_ilam_contains_all_articles():
    articles = [
        _article("기사1", "부정기사"),
        _article("기사2", "보류"),
        _article("기사3", "해당없음"),
    ]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    ws = wb["일람"]
    assert ws.max_row == 4  # header + 3


def test_ilam_header_columns_count():
    wb = _load(create_excel([], [], lang="ko"))
    ws = wb["일람"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 9) if ws.cell(row=1, column=c).value]
    assert len(headers) == 8


def test_ilam_article_link_in_correct_column():
    articles = [_article()]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    assert wb["일람"].cell(row=2, column=7).value == "https://example.com/news/1"


def test_ilam_article_title_in_correct_column():
    articles = [_article("특정기사제목")]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    assert wb["일람"].cell(row=2, column=6).value == "특정기사제목"


def test_ilam_datetime_formatted_correctly():
    articles = [_article()]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    assert wb["일람"].cell(row=2, column=3).value == "2024-03-25 09:30"


def test_ilam_no_published_at_shows_empty():
    article = _article()
    article["published_at"] = None
    wb = _load(create_excel([article], ["부정기사"], lang="ko"))
    # openpyxl returns None for empty cells (not empty string)
    assert not wb["일람"].cell(row=2, column=3).value


# ── category sheet routing ────────────────────────────────────────────────────

def test_category_sheet_only_contains_matching_articles():
    articles = [
        _article("부정1", "부정기사"),
        _article("중요1", "중요기사"),
        _article("부정2", "부정기사"),
    ]
    wb = _load(create_excel(articles, ["부정기사", "중요기사"], lang="ko"))
    assert wb["부정기사"].max_row == 3  # header + 2
    assert wb["중요기사"].max_row == 2  # header + 1


def test_holdup_sheet_contains_only_holdup_articles():
    articles = [
        _article("보류기사", "보류"),
        _article("일반기사", "부정기사"),
    ]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    assert wb["보류"].max_row == 2  # header + 1


def test_na_sheet_contains_only_na_articles():
    articles = [
        _article("무관련", "해당없음"),
        _article("일반", "부정기사"),
        _article("보류", "보류"),
    ]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    assert wb["해당없음"].max_row == 2  # header + 1


def test_empty_category_sheet_has_only_header():
    wb = _load(create_excel([], ["빈카테고리"], lang="ko"))
    assert wb["빈카테고리"].max_row == 1


# ── header styling ────────────────────────────────────────────────────────────

def test_ilam_header_is_bold():
    wb = _load(create_excel([], [], lang="ko"))
    assert wb["일람"].cell(row=1, column=1).font.bold is True


def test_ilam_header_has_background_fill():
    wb = _load(create_excel([], [], lang="ko"))
    cell = wb["일람"].cell(row=1, column=1)
    assert cell.fill.fgColor.rgb != "00000000"


def test_freeze_panes_set_on_ilam():
    wb = _load(create_excel([], [], lang="ko"))
    assert wb["일람"].freeze_panes == "A2"


# ── numbering ─────────────────────────────────────────────────────────────────

def test_ilam_row_numbers_are_sequential():
    articles = [_article(f"기사{i}") for i in range(5)]
    wb = _load(create_excel(articles, ["부정기사"], lang="ko"))
    ws = wb["일람"]
    for row in range(2, 7):
        assert ws.cell(row=row, column=1).value == row - 1


def test_category_sheet_row_numbers_restart_from_1():
    articles = [
        _article("기사1", "중요기사"),
        _article("기사2", "중요기사"),
    ]
    wb = _load(create_excel(articles, ["중요기사"], lang="ko"))
    ws = wb["중요기사"]
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=3, column=1).value == 2
