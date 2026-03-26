from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.i18n import get_strings

COL_WIDTHS_ILAM  = [6, 15, 20, 10, 15, 55, 45, 20]
COL_WIDTHS_OTHER = [6, 15, 20, 10, 15, 55, 45, 45]

HEADER_BG = "4472C4"
HEADER_FG = "FFFFFF"

# 데이터 필터링용 상수 (항상 한국어 고정)
_CAT_HOLDUP = "보류"
_CAT_NA = "해당없음"


def create_excel(
    articles: list[dict],
    category_names: list[str],
    lang: str = "ko",
) -> bytes:
    """
    시트 순서:
    1. 일람/一覧  - 모든 기사
    2. 사용자 정의 카테고리
    3. 보류/保留  - 분류 불가 기사
    """
    S = get_strings(lang)
    cols_ilam  = [S["col_no"], S["col_keyword"], S["col_datetime"], S["col_engine"],
                  S["col_media"], S["col_title"], S["col_link"], S["col_category"]]
    cols_other = [S["col_no"], S["col_keyword"], S["col_datetime"], S["col_engine"],
                  S["col_media"], S["col_title"], S["col_link"], S["col_reason"]]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. 일람/一覧 시트
    ws_ilam = wb.create_sheet(title=S["sheet_ilam"])
    _setup_header(ws_ilam, cols_ilam)
    for idx, article in enumerate(articles, 1):
        _add_row(ws_ilam, article, idx, is_ilam=True)
    _apply_column_widths(ws_ilam, is_ilam=True)

    # 2. 사용자 정의 카테고리 시트
    for cat_name in category_names:
        ws = wb.create_sheet(title=cat_name[:31])
        _setup_header(ws, cols_other)
        idx = 1
        for article in articles:
            if article.get("category") == cat_name:
                _add_row(ws, article, idx, is_ilam=False)
                idx += 1
        _apply_column_widths(ws, is_ilam=False)

    # 3. 보류/保留 시트
    ws_unc = wb.create_sheet(title=S["sheet_holdup"])
    _setup_header(ws_unc, cols_other)
    idx = 1
    for article in articles:
        if article.get("category") == _CAT_HOLDUP:
            _add_row(ws_unc, article, idx, is_ilam=False)
            idx += 1
    _apply_column_widths(ws_unc, is_ilam=False)

    # 4. 해당없음/該当なし 시트
    ws_na = wb.create_sheet(title=S["sheet_na"])
    _setup_header(ws_na, cols_other)
    idx = 1
    for article in articles:
        if article.get("category") == _CAT_NA:
            _add_row(ws_na, article, idx, is_ilam=False)
            idx += 1
    _apply_column_widths(ws_na, is_ilam=False)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _setup_header(ws, columns: list[str]) -> None:
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=HEADER_FG, size=11)

    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _add_row(ws, article: dict, row_num: int, is_ilam: bool) -> None:
    row = ws.max_row + 1

    pub_dt = article.get("published_at")
    date_str = pub_dt.strftime("%Y-%m-%d %H:%M") if isinstance(pub_dt, datetime) else ""

    last_col_value = article.get("category", "") if is_ilam else article.get("reason", "")

    values = [
        row_num,
        article.get("keyword", ""),
        date_str,
        article.get("search_engine", ""),
        article.get("source", ""),
        article.get("title", ""),
        article.get("link", ""),
        last_col_value,
    ]

    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        if col == 7 and value:  # 링크 컬럼
            try:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")
            except Exception:
                pass

    ws.row_dimensions[row].height = 40


def _apply_column_widths(ws, is_ilam: bool) -> None:
    widths = COL_WIDTHS_ILAM if is_ilam else COL_WIDTHS_OTHER
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
